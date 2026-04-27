import json
import os
import queue
import signal
import subprocess
import sys
import threading
import time
from collections import deque

import pandas as pd
from flask import Flask, jsonify, render_template, request, Response, send_from_directory
from openpyxl import Workbook, load_workbook

import convert_cookies


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, "URL_TIMMINGS.xlsx")
COOKIES_TXT = os.path.join(BASE_DIR, "cookies.txt")
COOKIES_JSON = os.path.join(BASE_DIR, "chrome_cookies.json")
PRIORITY_FILE = os.path.join(BASE_DIR, "priority.json")
CONFIG_FILE = os.path.join(BASE_DIR, "bot_config.json")
BOT_SCRIPT = os.path.join(BASE_DIR, "python_bot.py")

app = Flask(__name__, template_folder="templates", static_folder="static")


# ---------- bot process state ----------
class BotState:
    def __init__(self):
        self.proc: subprocess.Popen | None = None
        self.lock = threading.Lock()
        self.logs: deque[dict] = deque(maxlen=2000)
        self.cond = threading.Condition()
        self.seq = 0
        self.exit_code: int | None = None
        self.reader_thread: threading.Thread | None = None

    def append_log(self, line: str):
        with self.cond:
            self.seq += 1
            self.logs.append({"seq": self.seq, "line": line.rstrip("\n"), "ts": time.time()})
            self.cond.notify_all()

    def is_running(self) -> bool:
        return self.proc is not None and self.proc.poll() is None


bot = BotState()


def _reader(proc: subprocess.Popen):
    try:
        for line in iter(proc.stdout.readline, ""):
            if not line:
                break
            bot.append_log(line)
    finally:
        proc.stdout.close()
        rc = proc.wait()
        bot.exit_code = rc
        bot.append_log(f"--- bot exited with code {rc} ---")


# ---------- timings ----------
def read_timings() -> list[dict]:
    if not os.path.exists(EXCEL_FILE):
        return []
    df = pd.read_excel(EXCEL_FILE, header=1, usecols=[1, 2, 3, 4])
    df.columns = ["SUBJECT", "TYPE", "STRING", "URL LINK"]
    df = df.dropna(subset=["SUBJECT", "TYPE", "STRING", "URL LINK"])
    out = []
    for _, row in df.iterrows():
        out.append({
            "subject": str(row["SUBJECT"]).strip(),
            "type": str(row["TYPE"]).strip(),
            "string": str(row["STRING"]).strip(),
            "url": str(row["URL LINK"]).strip(),
        })
    return out


def write_timings(rows: list[dict]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["B1"], ws["C1"], ws["D1"], ws["E1"] = "SUBJECT", "TYPE", "STRING", "URL LINK"
    # Row 2 stays blank — that's what pandas reads as header (header=1, 0-indexed)
    for i, r in enumerate(rows, start=3):
        ws.cell(row=i, column=2, value=r.get("subject", "").strip())
        ws.cell(row=i, column=3, value=r.get("type", "").strip())
        ws.cell(row=i, column=4, value=r.get("string", "").strip())
        ws.cell(row=i, column=5, value=r.get("url", "").strip())
    wb.save(EXCEL_FILE)


# ---------- routes ----------
@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/timings", methods=["GET"])
def get_timings():
    try:
        return jsonify(read_timings())
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/timings", methods=["POST"])
def post_timings():
    rows = request.get_json(force=True)
    if not isinstance(rows, list):
        return jsonify({"error": "expected a list"}), 400
    write_timings(rows)
    return jsonify({"ok": True, "count": len(rows)})


@app.route("/api/priority", methods=["GET"])
def get_priority():
    if not os.path.exists(PRIORITY_FILE):
        return jsonify([])
    with open(PRIORITY_FILE) as f:
        return jsonify(json.load(f))


@app.route("/api/priority", methods=["POST"])
def post_priority():
    data = request.get_json(force=True)
    if not isinstance(data, list):
        return jsonify({"error": "expected a list"}), 400
    cleaned = [
        {"subject": str(e.get("subject", "")).strip(), "type": str(e.get("type", "")).strip()}
        for e in data
        if e.get("subject") and e.get("type")
    ]
    with open(PRIORITY_FILE, "w") as f:
        json.dump(cleaned, f, indent=2)
    return jsonify({"ok": True, "count": len(cleaned)})


@app.route("/api/config", methods=["GET"])
def get_config():
    if not os.path.exists(CONFIG_FILE):
        return jsonify({"headless": False, "slow_mo": 400})
    with open(CONFIG_FILE) as f:
        return jsonify(json.load(f))


@app.route("/api/config", methods=["POST"])
def post_config():
    data = request.get_json(force=True)
    cfg = {
        "headless": bool(data.get("headless", False)),
        "slow_mo": int(data.get("slow_mo", 400)),
    }
    with open(CONFIG_FILE, "w") as f:
        json.dump(cfg, f, indent=2)
    return jsonify({"ok": True, "config": cfg})


@app.route("/api/cookies", methods=["POST"])
def upload_cookies():
    if "file" not in request.files:
        return jsonify({"error": "no file"}), 400
    f = request.files["file"]
    f.save(COOKIES_TXT)
    try:
        n = convert_cookies.convert(COOKIES_TXT, COOKIES_JSON)
    except Exception as e:
        return jsonify({"error": f"conversion failed: {e}"}), 500
    return jsonify({"ok": True, "count": n})


@app.route("/api/cookies/status", methods=["GET"])
def cookies_status():
    out = {"cookies_txt": None, "chrome_cookies_json": None}
    if os.path.exists(COOKIES_TXT):
        out["cookies_txt"] = os.path.getmtime(COOKIES_TXT)
    if os.path.exists(COOKIES_JSON):
        out["chrome_cookies_json"] = os.path.getmtime(COOKIES_JSON)
    return jsonify(out)


@app.route("/api/bot/start", methods=["POST"])
def bot_start():
    with bot.lock:
        if bot.is_running():
            return jsonify({"error": "already running", "pid": bot.proc.pid}), 409
        if not os.path.exists(COOKIES_JSON):
            return jsonify({"error": "chrome_cookies.json missing — upload cookies first"}), 400
        if not os.path.exists(EXCEL_FILE):
            return jsonify({"error": "URL_TIMMINGS.xlsx missing"}), 400

        bot.exit_code = None
        bot.append_log(f"--- starting bot (pid will follow) ---")
        bot.proc = subprocess.Popen(
            [sys.executable, "-u", BOT_SCRIPT],
            cwd=BASE_DIR,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            bufsize=1,
            text=True,
        )
        bot.append_log(f"--- bot started, pid={bot.proc.pid} ---")
        bot.reader_thread = threading.Thread(target=_reader, args=(bot.proc,), daemon=True)
        bot.reader_thread.start()
        return jsonify({"ok": True, "pid": bot.proc.pid})


@app.route("/api/bot/stop", methods=["POST"])
def bot_stop():
    with bot.lock:
        if not bot.is_running():
            return jsonify({"error": "not running"}), 409
        proc = bot.proc
        try:
            proc.terminate()
        except Exception:
            pass

    # Wait outside the lock so /status etc. stay responsive.
    try:
        proc.wait(timeout=3)
    except subprocess.TimeoutExpired:
        try:
            proc.kill()
        except Exception:
            pass
    return jsonify({"ok": True})


@app.route("/api/bot/status", methods=["GET"])
def bot_status():
    return jsonify({
        "running": bot.is_running(),
        "pid": bot.proc.pid if bot.proc else None,
        "exit_code": bot.exit_code,
    })


@app.route("/api/bot/logs", methods=["GET"])
def bot_logs():
    """SSE stream. Optional ?since=<seq> resumes after a known sequence number."""
    try:
        since = int(request.args.get("since", "0"))
    except ValueError:
        since = 0

    def gen():
        # Replay any buffered lines newer than `since`.
        for entry in list(bot.logs):
            if entry["seq"] > since:
                yield f"data: {json.dumps(entry)}\n\n"
                since_local = entry["seq"]
        last_seq = since
        if bot.logs:
            last_seq = max(last_seq, bot.logs[-1]["seq"])

        while True:
            with bot.cond:
                bot.cond.wait(timeout=15)
                new = [e for e in bot.logs if e["seq"] > last_seq]
            if new:
                for entry in new:
                    yield f"data: {json.dumps(entry)}\n\n"
                last_seq = new[-1]["seq"]
            else:
                # heartbeat to keep connection alive
                yield ": ping\n\n"

    return Response(gen(), mimetype="text/event-stream", headers={
        "Cache-Control": "no-cache",
        "X-Accel-Buffering": "no",
    })


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=False, threaded=True)
